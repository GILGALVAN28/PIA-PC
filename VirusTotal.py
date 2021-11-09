import requests
import time
import json
import pandas
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import getpass
import os, sys


def VirusTotalUrl():

    #Generamos result.txt para guardar los resultados
    with open('result.txt', 'w') as file:
        file.write("")
    # leer urls sospechosas
    urls_sospechosas = []
    with open("urls_sospechosas.txt", "r") as file:
        for i in file:
            urls_sospechosas.append(i.strip("\n"))

    # api : 60f0dd4a80bc6228f3102ba63badf79348b578503a0f922a46a2eb1537e4e2b9
    API_key =  getpass.getpass('API key de Virus Total: ')
    url = 'https://www.virustotal.com/vtapi/v2/url/report'
    print("Cargando...")

    if os.path.exists("reporte_analizador_urls.xlsx"):

        os.remove("reporte_analizador_urls.xlsx")


    # numero de scans
    file = open("urls_sospechosas.txt", "r")
    line_count = 0
    for line in file:
        if line != "\n":
            line_count += 1
    file.close()

    # fecha del scan
    dt = datetime.now()
    dt = dt.strftime('%d-%m-%Y')

    # crear excel
    try:
        libro = load_workbook("reporte_analizador_urls.xlsx")
    except FileNotFoundError:
        libro = Workbook()

        # seleccionar hoja
    if len(libro.sheetnames) > 1:
        libro.active = 1
        hoja = libro.active
    else:
        hoja = libro.active
        hoja["A1"] = "Url"
        hoja["E1"] = 'Fecha de analisis'
        hoja["C1"] = "Total de análisis"
        hoja["D1"] = "Análisis positivos"
        hoja["B1"] = "Clasificación de riesgo"
        hoja["C2"] = line_count
        hoja["E2"] = dt
        libro.save("reporte_analizador_urls.xlsx")

    contador = 0

    try:
        for i in urls_sospechosas:
            parameters = {'apikey': API_key, 'resource': i}
            response = requests.get(url=url, params=parameters)
            json_response = json.loads(response.text)
            if json_response['response_code'] <= 0:
                with open('not Found result.txt', 'a') as notfound:
                    notfound.write("\tNOT found please Scan it manually\n")
            elif json_response['response_code'] >= 1:

                if json_response['positives'] <= 3:
                    with open('result.txt', 'a') as clean:
                        clean.write("\t Bajo \n")
                elif (json_response['positives'] > 3 and
                        json_response['positives'] <= 10):
                    with open('result.txt', 'a') as clean:
                        clean.write("\t Medio \n")
                        contador = contador + 1
                else:
                    with open('result.txt', 'a') as malicious:
                        malicious.write("\t Alto")
                        contador = contador+1
                time.sleep(15)
    except:
        
        print("Error:  Verifica tu API Key de Virus Total")
        exit()
    # guardar en excel
    result = open("result.txt", "r")
    result_url = result.readlines()
    for url in range(len(result_url)):
        hoja["B"+str(url + 2)] = str(result_url[url])
        libro.save("reporte_analizador_urls.xlsx")

    ligas = open("urls_sospechosas.txt", "r")
    result_ligas = ligas.readlines()
    for lig in range(len(result_ligas)):
        hoja["A"+str(lig + 2)] = str(result_ligas[lig])
        libro.save("reporte_analizador_urls.xlsx")

    try:
        libro = load_workbook("reporte_analizador_urls.xlsx")
    except FileNotFoundError:
        libro = Workbook()

        # seleccionar hoja
    if len(libro.sheetnames) > 1:
        libro.active = 1
        hoja = libro.active
    else:
        hoja = libro.active
        hoja["D2"] = contador
        libro.save("reporte_analizador_urls.xlsx")

    print("Se ha generado un reporte")
    with open('result.txt', 'w') as file:
        file.write("")
